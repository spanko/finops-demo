"""Excel workbook generator - one tab per recipe with full evidence."""

from __future__ import annotations

import json
from pathlib import Path

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from gt_finops.aggregate import summarize, all_findings


# GT brand colors for Excel formatting
PURPLE = "582C83"
PURPLE_SOFT = "E8E2F0"
ORANGE = "F26B23"
TEAL = "00838F"
WHITE = "FFFFFF"
LIGHT_GRAY = "FAFAFA"
BORDER_GRAY = "D1D1D1"

HEADER_FILL = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
HEADER_FONT = Font(color=WHITE, bold=True, name="Arial", size=11)
BODY_FONT = Font(color="1A1A1A", name="Arial", size=10)
MONEY_FONT = Font(color="1A1A1A", name="Arial", size=10, bold=True)
ALT_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

THIN = Side(border_style="thin", color=BORDER_GRAY)
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)


def write_excel_workbook(
    conn: duckdb.DuckDBPyConnection,
    output_path: Path,
    client_name: str,
) -> Path:
    """Write the findings evidence workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    # Summary tab
    _write_summary_sheet(wb, conn, client_name)

    # One tab per recipe with findings
    df_all = all_findings(conn)
    if not df_all.empty:
        for recipe_id, group in df_all.groupby("recipe_id"):
            sheet_name = f"{recipe_id}"[:31]  # Excel sheet name limit
            _write_recipe_sheet(wb, sheet_name, group, recipe_id)

    # Raw findings tab (for power users)
    if not df_all.empty:
        _write_raw_sheet(wb, df_all)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _write_summary_sheet(wb: Workbook, conn, client_name: str) -> None:
    ws = wb.create_sheet("Summary")
    summary = summarize(conn)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    ws["A1"] = "Grant Thornton · FinOps Findings"
    ws["A1"].font = Font(color=PURPLE, bold=True, size=18, name="Arial")
    ws["A2"] = f"Client: {client_name}"
    ws["A2"].font = Font(color="333333", size=12, name="Arial")

    # Headline stats
    ws["A4"] = "Total findings"
    ws["B4"] = summary.total_findings
    ws["A5"] = "Gross annual savings"
    ws["B5"] = summary.total_gross_usd
    ws["B5"].number_format = '"$"#,##0'
    ws["A6"] = "Capturable annual savings"
    ws["B6"] = summary.total_capturable_usd
    ws["B6"].number_format = '"$"#,##0'
    ws["B6"].font = Font(bold=True, color=ORANGE, size=12, name="Arial")

    for row in range(4, 7):
        ws[f"A{row}"].font = Font(bold=True, color="333333", name="Arial", size=10)

    # Category breakdown
    ws["A8"] = "By category"
    ws["A8"].font = Font(bold=True, color=PURPLE, size=12, name="Arial")

    headers = ["Category", "Findings", "Gross annual", "Capturable annual"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=9, column=i + 1, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    row = 10
    for cat, stats in summary.by_category.items():
        ws.cell(row=row, column=1, value=cat).font = BODY_FONT
        ws.cell(row=row, column=2, value=stats["count"]).font = BODY_FONT
        c3 = ws.cell(row=row, column=3, value=stats["gross"])
        c3.number_format = '"$"#,##0'; c3.font = BODY_FONT
        c4 = ws.cell(row=row, column=4, value=stats["capturable"])
        c4.number_format = '"$"#,##0'; c4.font = MONEY_FONT
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = BORDER
            if row % 2 == 0:
                ws.cell(row=row, column=col).fill = ALT_FILL
        row += 1

    # Per-recipe breakdown
    row += 2
    ws.cell(row=row, column=1, value="By recipe").font = Font(
        bold=True, color=PURPLE, size=12, name="Arial"
    )
    row += 1
    headers = ["Recipe", "Name", "Findings", "Capturable annual"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    row += 1

    for rid, stats in summary.by_recipe.items():
        ws.cell(row=row, column=1, value=rid).font = BODY_FONT
        ws.cell(row=row, column=2, value=stats["name"]).font = BODY_FONT
        ws.cell(row=row, column=3, value=stats["count"]).font = BODY_FONT
        c = ws.cell(row=row, column=4, value=stats["capturable"])
        c.number_format = '"$"#,##0'; c.font = MONEY_FONT
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    ws.freeze_panes = "A3"


def _write_recipe_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, recipe_id: str) -> None:
    ws = wb.create_sheet(sheet_name)

    # Title row
    recipe_name = df["recipe_name"].iloc[0] if len(df) > 0 else recipe_id
    ws["A1"] = f"Recipe {recipe_id}: {recipe_name}"
    ws["A1"].font = Font(bold=True, color=PURPLE, size=14, name="Arial")
    ws.merge_cells("A1:H1")

    # Columns to expose
    columns = [
        ("entity_name", "Entity", 30),
        ("entity_type", "Type", 14),
        ("current_state", "Current state", 40),
        ("recommended_state", "Recommendation", 40),
        ("gross_annual_savings_usd", "Gross USD/yr", 13),
        ("capturable_annual_savings_usd", "Capturable USD/yr", 16),
        ("confidence", "Confidence", 12),
        ("risk_level", "Risk", 10),
        ("days_to_capture", "Days", 7),
        ("suggested_owner", "Owner", 24),
        ("dependencies", "Dependencies", 40),
        ("evidence", "Evidence", 45),
    ]

    for idx, (_, label, width) in enumerate(columns):
        col_letter = get_column_letter(idx + 1)
        ws.column_dimensions[col_letter].width = width
        cell = ws.cell(row=3, column=idx + 1, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    # Rows
    for row_idx, (_, finding) in enumerate(df.iterrows(), start=4):
        for col_idx, (col_name, _, _) in enumerate(columns, start=1):
            val = finding.get(col_name)
            # Nice formatting for JSON fields
            if col_name in ("dependencies", "evidence") and isinstance(val, str):
                try:
                    parsed = json.loads(val)
                    if isinstance(parsed, list):
                        val = "\n".join(f"• {x}" for x in parsed)
                    elif isinstance(parsed, dict):
                        val = "\n".join(f"{k}: {v}" for k, v in parsed.items())
                except (json.JSONDecodeError, TypeError):
                    pass

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = LEFT
            cell.border = BORDER

            if col_name in ("gross_annual_savings_usd", "capturable_annual_savings_usd"):
                cell.number_format = '"$"#,##0'
                cell.alignment = RIGHT
                if col_name == "capturable_annual_savings_usd":
                    cell.font = MONEY_FONT
            if col_name == "days_to_capture":
                cell.alignment = CENTER

            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

        ws.row_dimensions[row_idx].height = 42

    ws.freeze_panes = "A4"


def _write_raw_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Write a raw dump tab for power users."""
    ws = wb.create_sheet("_raw")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(df.columns, start=1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            elif hasattr(val, "isoformat"):
                val = val.isoformat()
            ws.cell(row=row_idx, column=col_idx, value=val).font = BODY_FONT

    ws.freeze_panes = "A2"
